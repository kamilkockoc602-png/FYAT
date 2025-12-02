from flask import Flask, request, jsonify, send_from_directory
from openpyxl import load_workbook
import os
from datetime import datetime

app = Flask(__name__, static_folder='.')

@app.route('/upload_excel', methods=['POST'])
def upload_excel():
    file = request.files.get('excel')
    username = request.form.get('username', 'guest')
    if not file:
        return jsonify({'error':'Dosya gerekli'}), 400
    wb = load_workbook(file.stream, data_only=True)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        origin, destination, price = row
        if origin and destination and price is not None:
            data.append({
                'origin': origin,
                'destination': destination,
                'price': price,
                'username': username,
                'uploaded_at': datetime.now().isoformat()
            })
    return jsonify(data)

# Serve frontend
@app.route('/', defaults={'path':'flixbus_frontend.html'})
@app.route('/<path:path>')
def serve(path):
    if os.path.exists(path):
        return send_from_directory('.', path)
    return 'Not Found', 404

if __name__ == '__main__':
    app.run(debug=True)
