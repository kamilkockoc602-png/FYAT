from flask import Flask, request, jsonify, send_from_directory
from openpyxl import load_workbook
import os, json
from datetime import datetime

app = Flask(__name__, static_folder='.')

BAKANLIK_FILE = 'bakanlik_data.json'
# ADMIN_KEY environment variable ile kontrol edin; yoksa default kullanılır (local test için)
ADMIN_KEY = os.environ.get('ADMIN_KEY', 'dev_admin_key_123')

def save_json_file(path, data):
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def load_json_file(path):
    if not os.path.exists(path):
        return []
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)

# Yeni endpoint: bakanlik excel yükle
@app.route('/upload_bakanlik_excel', methods=['POST'])
def upload_bakanlik_excel():
    # Basit admin key kontrolü
    admin_key = request.form.get('adminKey') or request.headers.get('X-Admin-Key')
    if admin_key != ADMIN_KEY:
        return jsonify({'error': 'Unauthorized'}), 403

    file = request.files.get('file') or request.files.get('excel') or request.files.get('xlsx')
    if not file:
        return jsonify({'error': 'Dosya gerekli (form field: file)'}), 400

    try:
        wb = load_workbook(file.stream, data_only=True)
    except Exception as e:
        return jsonify({'error': 'Excel okunamadı', 'detail': str(e)}), 400

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        return jsonify({'error': 'Excel içinde veri bulunamadı veya başlık satırı eksik'}), 400

    # İlk satırı başlık olarak al
    headers = [str(h).strip().lower() if h is not None else '' for h in rows[0]]

    # Map edilecek alan isimleri (türkçe ve ingilizce varyantlar)
    def map_header_name(h):
        h = h.lower()
        if any(k in h for k in ['route','güzergah','kalkış','kalkis','kalkis-varis','kalkış-varış']):
            return 'route'
        if any(k in h for k in ['price','fiyat','tarife']):
            return 'price'
        if any(k in h for k in ['discount','ind','indirim','discounted']):
            return 'discounted'
        if any(k in h for k in ['km']):
            return 'km'
        if any(k in h for k in ['unit','birim']):
            return 'unit'
        return h

    mapped = [map_header_name(h) for h in headers]

    out = []
    for r in rows[1:]:
        if all([c is None or (isinstance(c, str) and c.strip()=='') for c in r]):
            continue
        obj = {}
        for i, cell in enumerate(r):
            key = mapped[i] if i < len(mapped) else f'col{i}'
            if cell is None:
                value = ''
            else:
                value = cell
            obj[key] = value
        # normalize: ensure route and price exist
        if not obj.get('route'):
            # attempt to combine origin/destination if that format used in Excel
            if obj.get('origin') and obj.get('destination'):
                obj['route'] = f"{obj.get('origin')} - {obj.get('destination')}"
            else:
                # skip rows missing route
                continue
        # convert price to number if possible
        price_val = obj.get('price', '')
        try:
            if isinstance(price_val, str):
                pv = price_val.replace('₺','').replace(',','').strip()
                obj['price'] = float(pv) if pv != '' else None
            elif isinstance(price_val, (int,float)):
                obj['price'] = float(price_val)
            else:
                obj['price'] = None
        except:
            obj['price'] = None

        # add metadata
        obj['uploaded_at'] = datetime.now().isoformat()
        out.append({
            'route': str(obj.get('route')).strip(),
            'price': obj.get('price'),
            'discounted': obj.get('discounted', ''),
            'km': obj.get('km', ''),
            'unit': obj.get('unit',''),
            'meta': {k:v for k,v in obj.items() if k not in ['route','price','discounted','km','unit','uploaded_at']},
            'uploaded_at': obj['uploaded_at']
        })

    if not out:
        return jsonify({'error':'Excel içinde uygun veri bulunamadı'}), 400

    
    save_json_file(BAKANLIK_FILE, out)

    return jsonify({'ok': True, 'count': len(out), 'file': BAKANLIK_FILE})

# mevcut serve route'larınız...
@app.route('/', defaults={'path':''})
@app.route('/<path:path>')
def serve(path):
    if path and os.path.exists(path):
        return send_from_directory('.', path)
    return send_from_directory('.', 'index.html')

if __name__ == '__main__':
    # debug veya production ayarlarınıza göre
    app.run(debug=True)