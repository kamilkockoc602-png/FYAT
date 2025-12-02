from flask import Flask, request, send_file, jsonify
from openpyxl import load_workbook
from PIL import Image
import pytesseract
import io
import re
import datetime
from pdf2image import convert_from_bytes

app = Flask(__name__)

# If Tesseract is not in PATH, uncomment and set:
# pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

def ocr_from_image(image_bytes, lang):
    image = Image.open(io.BytesIO(image_bytes))
    return pytesseract.image_to_string(image, lang=lang)

def ocr_from_pdf(pdf_bytes, lang):
    images = convert_from_bytes(pdf_bytes)
    text = ''
    for img in images:
        text += pytesseract.image_to_string(img, lang=lang) + '\n'
    return text

@app.route('/upload', methods=['POST'])
def upload():
    files = request.files.getlist('files')
    lang = request.form.get('lang', 'tur')
    pattern = r"""([A-ZÇĞİÖŞÜa-zçğıöşü\s]+)\s*-\s*([A-ZÇĞİÖŞÜa-zçğıöşü\s]+)\s*=\s*(\d+)"""
    results = []
    for file in files:
        content = file.read()
        if file.filename.lower().endswith('.pdf'):
            text = ocr_from_pdf(content, lang)
        else:
            text = ocr_from_image(content, lang)
        for kalkis, varis, fiyat in re.findall(pattern, text):
            results.append({
                'kalkis': kalkis.strip(),
                'varis': varis.strip(),
                'fiyat': fiyat.strip(),
                'tarih': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            })
    return jsonify(results)

@app.route('/generate', methods=['POST'])
def generate():
    data = request.json.get('data', [])
    wb = load_workbook('Yeni_Fiyat_Talep.xlsx')
    ws = wb.active
    row = ws.max_row + 1
    for item in data:
        ws[f"B{row}"] = item['kalkis']
        ws[f"C{row}"] = item['varis']
        ws[f"H{row}"] = int(item['fiyat'])
        ws[f"I{row}"] = item['tarih']
        row += 1
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='Fiyat_Talep_Sonuclu.xlsx')

if __name__ == '__main__':
    app.run(debug=True)
