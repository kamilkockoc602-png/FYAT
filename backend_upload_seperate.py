from flask import Flask, request, jsonify, send_from_directory
from openpyxl import load_workbook
from flask_cors import CORS
import os, json, uuid
from datetime import datetime

app = Flask(__name__, static_folder='.')
CORS(app)  # üretimde origins ile kısıtlayın

# Mevcut resmi tarifeler ayrı tutuluyor; yeni yüklemeler uploads.json içinde saklanır
BAKANLIK_FILE = 'bakanlik_data.json'
UPLOADS_FILE = 'uploads.json'
ADMIN_KEY = os.environ.get('ADMIN_KEY', 'dev_admin_key_123')  # opsiyonel admin key

def load_json(path):
    if not os.path.exists(path):
        return []
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_json(path, data):
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def is_admin(req):
    # Admin kontrolü (opsiyonel): X-Admin-Key veya Authorization: Bearer <ADMIN_KEY>
    k = req.headers.get('X-Admin-Key') or ''
    auth = req.headers.get('Authorization') or ''
    if k and k == ADMIN_KEY: return True
    if auth.startswith('Bearer ') and auth.split(' ',1)[1] == ADMIN_KEY:
        return True
    return False

@app.route('/api/upload_bakanlik_excel', methods=['POST'])
def upload_bakanlik_excel():
    """
    Excel yüklemesini UPLOADS_FILE'a ekler (BAKANLIK_FILE'a dokunmaz).
    X-User header (veya form field username) ile kimin yüklediği kaydedilir.
    """
    uploader = request.headers.get('X-User') or request.form.get('username') or 'anonymous'
    file = request.files.get('file') or request.files.get('excel')
    if not file:
        return jsonify({'error': 'Dosya gerekli (form field: file veya excel)'}), 400

    try:
        wb = load_workbook(file.stream, data_only=True)
    except Exception as e:
        return jsonify({'error':'Excel okunamadı', 'detail': str(e)}), 400

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        return jsonify({'error':'Excel içinde veri bulunamadı veya başlık satırı eksik'}), 400

    headers = [str(h).strip().lower() if h is not None else '' for h in rows[0]]

    def map_header(h):
        h = h.lower()
        if any(k in h for k in ['origin','kalk','kalkış','kalkis','from']): return 'origin'
        if any(k in h for k in ['destination','varış','varis','to']): return 'destination'
        if any(k in h for k in ['price','fiyat','tarife']): return 'price'
        return h

    mapped = [map_header(h) for h in headers]

    uploads = load_json(UPLOADS_FILE)
    new_items = []
    for r in rows[1:]:
        if all(c is None or (isinstance(c, str) and c.strip()=='') for c in r):
            continue
        obj = {}
        for i, cell in enumerate(r):
            key = mapped[i] if i < len(mapped) else f'col{i}'
            obj[key] = cell
        origin = obj.get('origin') or obj.get('kalkis') or ''
        destination = obj.get('destination') or obj.get('varis') or ''
        route = f"{str(origin).strip()} - {str(destination).strip()}" if origin or destination else (obj.get('route') or '')
        # price parse
        price_raw = obj.get('price', None)
        price = None
        try:
            if price_raw is None or price_raw == '':
                price = None
            elif isinstance(price_raw, (int, float)):
                price = float(price_raw)
            else:
                pv = str(price_raw).replace('₺','').replace(',','').strip()
                price = float(pv) if pv != '' else None
        except:
            price = None

        item = {
            'id': uuid.uuid4().hex,
            'route': route,
            'price': price,
            'meta': {k:v for k,v in obj.items() if k not in ['origin','destination','price','route']},
            'uploaded_by': uploader,
            'uploaded_at': datetime.utcnow().isoformat()
        }
        new_items.append(item)
        uploads.append(item)

    if not new_items:
        return jsonify({'error':'Excel içinde uygun veri bulunamadı veya boş satırlar'}), 400

    save_json(UPLOADS_FILE, uploads)
    return jsonify({'ok': True, 'inserted': len(new_items), 'ids': [it['id'] for it in new_items]})

@app.route('/api/uploads', methods=['GET'])
def list_uploads():
    """
    - Eğer istek admin (ADMIN_KEY) ile yapılırsa tüm kayıtları döner.
    - Eğer ?user=username parametresi veya X-User header verildiyse sadece o kullanıcıya ait kayıtları döner.
    - Aksi halde 401 döner.
    """
    uploads = load_json(UPLOADS_FILE)
    # admin ile tam liste
    if is_admin(request):
        return jsonify(uploads)

    # kendi kayıtlarını al
    user_q = request.args.get('user') or request.headers.get('X-User')
    if user_q:
        user_q = user_q.strip()
        my = [u for u in uploads if u.get('uploaded_by') == user_q]
        return jsonify(my)
    # yetkisiz genel liste istemiyoruz
    return jsonify({'error':'Unauthorized or specify ?user=username / X-User header'}), 401

@app.route('/api/uploads/<id>', methods=['DELETE'])
def delete_upload(id):
    """
    Silme izni:
      - admin (ADMIN_KEY) ise silebilir; veya
      - istek yapan X-User header değeri, silinmek istenen kaydı yükleyen (uploaded_by) ile aynı ise silebilir.
    """
    uploads = load_json(UPLOADS_FILE)
    target = next((u for u in uploads if u.get('id') == id), None)
    if not target:
        return jsonify({'error':'Not found'}), 404

    # admin kontrol
    if is_admin(request):
        new = [d for d in uploads if d.get('id') != id]
        save_json(UPLOADS_FILE, new)
        return jsonify({'ok': True, 'deleted': id})

    # kullanıcı kontrolü
    requester = request.headers.get('X-User') or request.args.get('user')
    if not requester:
        return jsonify({'error':'Unauthorized - no X-User header provided'}), 401

    if requester != target.get('uploaded_by'):
        return jsonify({'error':'Unauthorized - you can only delete your own uploads'}), 401

    # sil
    new = [d for d in uploads if d.get('id') != id]
    save_json(UPLOADS_FILE, new)
    return jsonify({'ok': True, 'deleted': id})

# Statik dosya serve (opsiyonel)
@app.route('/', defaults={'path':''})
@app.route('/<path:path>')
def serve(path):
    if path and os.path.exists(path):
        return send_from_directory('.', path)
    return send_from_directory('.', 'index.html')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)), debug=True)